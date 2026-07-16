import json
import os
from io import BytesIO

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo

_INVALID_SHEET_CHARS = (":", "\\", "/", "?", "*", "[", "]")

_CONTENT_TYPE_XLSX = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_HEADER_FILL = PatternFill(
    start_color="0A2540", end_color="0A2540", fill_type="solid"
)
_HEADER_FONT = Font(
    name="Calibri", size=11, bold=True, color="FFFFFF"
)
_TITLE_FONT = Font(
    name="Calibri", size=16, bold=True, color="0A2540"
)
_SUBTITLE_FONT = Font(
    name="Calibri", size=11, bold=True, color="1F2937"
)
_LABEL_FONT = Font(
    name="Calibri", size=10, bold=True, color="374151"
)
_TEXT_FONT = Font(
    name="Calibri", size=10, color="111827"
)
_SECTION_FILL = PatternFill(
    start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"
)
_SECTION_FONT = Font(
    name="Calibri", size=11, bold=True, color="0A2540"
)
_SUMMARY_TITLE_FILL = PatternFill(
    start_color="EAF1F5", end_color="EAF1F5", fill_type="solid"
)
_SUMMARY_TITLE_FONT = Font(
    name="Calibri", size=11, bold=True, color="1F5C7A"
)
_ZEBRA_FILL = PatternFill(
    start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
)

_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(
    left=_THIN, right=_THIN, top=_THIN, bottom=_THIN
)

_CENTER = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
_LEFT = Alignment(
    horizontal="left", vertical="center", wrap_text=True
)
_LEFT_TOP = Alignment(
    horizontal="left", vertical="top", wrap_text=True
)
_RIGHT = Alignment(
    horizontal="right", vertical="center", wrap_text=True
)

_FOOTER_TEXT = "ASIMS – Rapport généré automatiquement"

_LOGO_CANDIDATES = (
    "images/logo.png",
    "static/images/logo.png",
)


def _sanitize_sheet_name(name, fallback_index):

    name = str(name).strip()

    for char in _INVALID_SHEET_CHARS:
        name = name.replace(char, " ")

    name = name[:31].strip() or "Feuille{0}".format(fallback_index)

    return name


def _is_rich(value):

    return isinstance(value, CellRichText)


def _stringify(value):

    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except TypeError:
            return str(value)

    return str(value)


def _rows_from_value(value):

    if value is None or isinstance(value, (str, int, float, bool)):
        return [[_stringify(value)]]

    if isinstance(value, dict):
        return [[str(key), _stringify(sub)] for key, sub in value.items()]

    if isinstance(value, list):

        if value and all(isinstance(item, dict) for item in value):

            keys = []

            for item in value:
                for key in item.keys():
                    if key not in keys:
                        keys.append(key)

            rows = [[str(key) for key in keys]]

            for item in value:
                rows.append([_stringify(item.get(key)) for key in keys])

            return rows

        return [[_stringify(item)] for item in value]

    return [[_stringify(value)]]


_REPORT_TITLE = "Rapport statistique des GAB"

_KPI_LABELS = (
    ("total_gab", "Total GAB"),
    ("operational_gab", "GAB opérationnels"),
    ("critical_gab", "GAB critiques"),
    ("offline_gab", "GAB hors service"),
    ("passive_gab", "GAB passifs"),
    ("total_incidents", "Total incidents"),
    ("total_interventions", "Total interventions"),
    ("total_filiales", "Total filiales"),
    ("total_fournisseurs", "Total fournisseurs"),
    ("availability", "Disponibilité (%)"),
)


def _build_executive_summary_lines(statistics):

    groups = []

    kpis = statistics.get("kpis") or {}
    sla = statistics.get("sla") or {}
    analytics = statistics.get("analytics") or {}
    health = statistics.get("health") or []
    recommendations = statistics.get("recommendations") or {}

    availability = kpis.get("availability")
    total_gab = kpis.get("total_gab")
    operational = kpis.get("operational_gab")
    critical = kpis.get("critical_gab")
    offline = kpis.get("offline_gab")
    total_incidents = kpis.get("total_incidents")
    total_interventions = kpis.get("total_interventions")
    sla_rate = sla.get("sla_rate")
    closed = sla.get("closed_incidents")
    mttr = analytics.get("mttr")

    disponibilite = []
    if availability is not None:
        disponibilite.append(("Disponibilité globale", availability))
    if total_gab is not None:
        disponibilite.append(("Parc total", "{0} GAB".format(total_gab)))
    if operational is not None:
        disponibilite.append(("GAB opérationnels", operational))
    if critical is not None:
        disponibilite.append(("GAB critiques", critical))
    if offline is not None:
        disponibilite.append(("GAB hors service", offline))

    incidents = []
    if total_incidents is not None:
        incidents.append(("Incidents enregistrés", total_incidents))
    if closed is not None:
        incidents.append(("Incidents clôturés", closed))
    if total_interventions is not None:
        incidents.append(("Interventions réalisées", total_interventions))

    performance = []
    if sla_rate is not None:
        performance.append(("Taux de respect SLA", "{0} %".format(sla_rate)))
    if mttr is not None:
        performance.append(("MTTR moyen", mttr))

    sante = []
    if health:
        critical_atms = [
            item for item in health
            if item.get("status") == "Critique"
        ]
        sante.append((
            "GAB en état critique",
            "{0} sur {1}".format(len(critical_atms), len(health)),
        ))

    recommandations = []
    if recommendations:
        for item in recommendations:
            title = item.get("title")
            if title:
                recommandations.append(("Recommandation", title))

    if disponibilite:
        groups.append(("Disponibilité", disponibilite))
    if incidents:
        groups.append(("Incidents", incidents))
    if performance:
        groups.append(("Performance", performance))
    if sante:
        groups.append(("Santé du réseau", sante))
    if recommandations:
        groups.append(("Recommandations", recommandations))

    if not groups:
        groups.append(
            (
                "Résumé",
                [("Information", "Aucune donnée statistique disponible "
                  "pour générer le résumé.")],
            )
        )

    return groups


def _build_executive_summary_rows(statistics, filters, user):

    rows = []

    rows.append(["", ""])
    rows.append(["", ""])
    rows.append(["", ""])
    rows.append(["", ""])
    rows.append(["", ""])
    rows.append(["", ""])

    generated_at = timezone.now().strftime("%d/%m/%Y %H:%M:%S")
    if user and getattr(user, "is_authenticated", False):
        username = getattr(user, "get_full_name", lambda: "")() or getattr(
            user, "username", ""
        )
        user_label = str(username) if username else str(
            getattr(user, "username", "")
        )
    else:
        user_label = "(utilisateur non connecté)"

    rows.append(["Rapport", _REPORT_TITLE])
    rows.append(["Généré le", generated_at])
    rows.append(["Utilisateur", user_label])

    rows.append([])
    if filters:
        rows.append(["Filtres appliqués"])
        for key, value in filters.items():
            rows.append([str(key), _stringify(value)])

    rows.append([])
    rows.append(["KPIs principaux"])
    kpis = statistics.get("kpis") or {}
    for key, label in _KPI_LABELS:
        rows.append([label, _stringify(kpis.get(key))])

    rows.append([])
    rows.append(["Résumé exécutif"])
    for title, facts in _build_executive_summary_lines(statistics):
        rows.append([title])
        for label, value in facts:
            rows.append(["•", label, _stringify(value)])
        rows.append([])

    return rows


def _build_kpis_rows(statistics, filters):

    rows = []

    rows.append(["Indicateur", "Valeur"])

    kpis = statistics.get("kpis") or {}
    for key, label in _KPI_LABELS:
        rows.append([label, _stringify(kpis.get(key))])

    return rows


_RANKING_SECTIONS = (
    ("gabs", "Top GAB"),
    ("fournisseurs", "Top Fournisseurs"),
    ("agences", "Top Agences"),
    ("villes", "Top Villes"),
)


def _build_ranking_rows(statistics, filters):

    rows = []

    rankings = statistics.get("rankings") or {}

    for key, title in _RANKING_SECTIONS:

        rows.append([title])

        data = rankings.get(key)

        if data:
            rows.extend(_rows_from_value(data))
        else:
            rows.append(["(aucune donnée)"])

        rows.append([])

    return rows


_DISTRIBUTION_SECTIONS = (
    ("etat_distribution", "Distribution par etat"),
    ("ville_distribution", "Distribution par ville"),
    ("type_distribution", "Distribution par type"),
    ("agence_distribution", "Distribution par agence"),
    ("incidents_categorie", "Incidents par categorie"),
    ("incidents_fournisseur", "Incidents par fournisseur"),
    ("incidents_region", "Incidents par region"),
    ("incidents_filiale", "Incidents par filiale"),
    ("interventions_responsable", "Interventions par responsable"),
)


def _build_distribution_rows(statistics, filters):

    rows = []

    distribution = statistics.get("distribution") or {}

    for key, title in _DISTRIBUTION_SECTIONS:

        rows.append([title])

        data = distribution.get(key)

        if data:
            rows.extend(_rows_from_value(data))
        else:
            rows.append(["(aucune donnée)"])

        rows.append([])

    return rows


_EVOLUTION_SECTIONS = (
    ("incidents_month", "Evolution des incidents par mois"),
    ("interventions_month", "Evolution des interventions par mois"),
)


def _build_evolution_rows(statistics, filters):

    rows = []

    evolution = statistics.get("evolution") or {}

    for key, title in _EVOLUTION_SECTIONS:

        rows.append([title])

        data = evolution.get(key)

        if data:
            rows.extend(_rows_from_value(data))
        else:
            rows.append(["(aucune donnée)"])

        rows.append([])

    return rows


def _build_health_rows(statistics, filters):

    rows = []

    health = statistics.get("health") or []

    if health:
        rows.extend(_rows_from_value(health))
    else:
        rows.append(["(aucune donnée)"])

    return rows


def _build_recommendations_rows(statistics, filters):

    rows = []

    recommendations = statistics.get("recommendations") or []

    if recommendations:
        rows.extend(_rows_from_value(recommendations))
    else:
        rows.append(["(aucune donnée)"])

    return rows


def _build_filters_rows(statistics, filters):

    return _rows_from_value(filters or {})


_SHEET_BUILDERS = (
    ("Résumé Exécutif", _build_executive_summary_rows, None),
    ("KPIs", _build_kpis_rows, "kpis"),
    ("Ranking", _build_ranking_rows, "rankings"),
    ("Filtres", _build_filters_rows, None),
    ("Distribution", _build_distribution_rows, "distribution"),
    ("Health", _build_health_rows, "health"),
    ("Evolution", _build_evolution_rows, "evolution"),
    ("Recommendations", _build_recommendations_rows, "recommendations"),
)


def _find_logo_path():

    candidates = []

    if getattr(settings, "BASE_DIR", None) is not None:
        base = settings.BASE_DIR
        candidates.append(os.path.join(str(base), "static", "images", "logo.png"))

    here = os.path.dirname(os.path.abspath(__file__))
    candidates.append(
        os.path.normpath(
            os.path.join(here, "..", "..", "static", "images", "logo.png")
        )
    )
    candidates.append(
        os.path.normpath(os.path.join(here, "..", "static", "images", "logo.png"))
    )

    seen = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        if os.path.isfile(candidate):
            return candidate

    return None


def _add_logo(worksheet, logo_path):

    try:
        img = XLImage(logo_path)
    except Exception:
        return

    max_width = 140
    ratio = img.height / float(img.width) if img.width else 1
    img.width = max_width
    img.height = int(max_width * ratio)
    img.anchor = "A1"
    worksheet.add_image(img)


def _style_header_row(worksheet, row_index, n_cols):

    for col in range(1, n_cols + 1):
        cell = worksheet.cell(row=row_index, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _apply_table_style(worksheet, first_row, last_row, n_cols):

    if last_row < first_row or n_cols < 1:
        return

    last_col_letter = get_column_letter(n_cols)

    for row in range(first_row, last_row + 1):
        for col in range(1, n_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = _BORDER
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = _LEFT
            if cell.font is None or cell.font.size is None:
                cell.font = _TEXT_FONT

    for row in range(first_row + 1, last_row + 1, 2):
        for col in range(1, n_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = _ZEBRA_FILL


def _add_footer(worksheet, last_row):

    footer_row = last_row + 2
    worksheet.cell(row=footer_row, column=1, value=_FOOTER_TEXT)
    footer_cell = worksheet.cell(row=footer_row, column=1)
    footer_cell.font = Font(
        name="Calibri", size=9, italic=True, color="9CA3AF"
    )


def _auto_width(worksheet, max_width=60):

    widths = {}

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column
            if col not in widths or length > widths[col]:
                widths[col] = length

    for col, length in widths.items():
        worksheet.column_dimensions[get_column_letter(col)].width = min(
            max(length + 3, 12), max_width
        )


def _build_workbook(sheets, user=None):

    workbook = Workbook()
    workbook.remove(workbook.active)

    logo_path = _find_logo_path()

    for index, (name, rows) in enumerate(sheets, start=1):

        worksheet = workbook.create_sheet(
            title=_sanitize_sheet_name(name, index)
        )

        is_summary = name == "Résumé Exécutif"

        if is_summary and logo_path:
            _add_logo(worksheet, logo_path)

        for row in rows:
            worksheet.append([_stringify(cell) for cell in row])

        if worksheet.max_row == 0:
            continue

        if is_summary:
            _style_summary_sheet(worksheet)
        else:
            _style_data_sheet(worksheet)

        if name == "KPIs":
            _apply_kpi_conditional_formatting(worksheet)

        _auto_width(worksheet)
        _add_footer(worksheet, worksheet.max_row)

    return workbook


def _apply_kpi_conditional_formatting(worksheet):

    for row in range(1, worksheet.max_row + 1):
        label = worksheet.cell(row=row, column=1).value
        if not isinstance(label, str):
            continue

        value_cell = worksheet.cell(row=row, column=2)
        rule = None

        if label == "Disponibilité (%)":
            rule = CellIsRule(
                operator="lessThan",
                formula=["95"],
                fill=PatternFill(
                    start_color="FEE2E2", end_color="FEE2E2",
                    fill_type="solid"
                ),
                font=Font(color="B91C1C", bold=True),
            )
        elif label in (
            "GAB critiques", "GAB hors service", "Total incidents"
        ):
            rule = CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill(
                    start_color="FEF3C7", end_color="FEF3C7",
                    fill_type="solid"
                ),
                font=Font(color="92400E", bold=True),
            )

        if rule is not None:
            worksheet.conditional_formatting.add(value_cell.coordinate, rule)


def _style_summary_sheet(worksheet):

    last_row = worksheet.max_row

    worksheet.cell(row=7, column=2).font = _TITLE_FONT
    worksheet.cell(row=8, column=2).font = _SUBTITLE_FONT
    worksheet.cell(row=9, column=2).font = _SUBTITLE_FONT
    worksheet.cell(row=10, column=2).font = _TEXT_FONT

    data_start = 12
    for row in range(data_start, last_row + 1):
        first = worksheet.cell(row=row, column=1)
        second = worksheet.cell(row=row, column=2)

        if first.value in (None, ""):
            continue

        if second.value in (None, ""):
            first.font = _SECTION_FONT
            first.fill = _SECTION_FILL
            worksheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=2
            )
            first.alignment = _LEFT
            continue

        first.font = _LABEL_FONT
        second.font = _TEXT_FONT
        first.alignment = _LEFT_TOP
        second.alignment = _LEFT_TOP

    worksheet.freeze_panes = "A11"

    summary_titles = {
        "Disponibilité", "Incidents", "Performance",
        "Santé du réseau", "Recommandations", "Résumé",
    }
    summary_start = None
    for row in range(data_start, last_row + 1):
        if worksheet.cell(row=row, column=1).value == "Résumé exécutif":
            summary_start = row + 1
            break

    if summary_start is not None:
        for row in range(summary_start, last_row + 1):
            first = worksheet.cell(row=row, column=1)
            second = worksheet.cell(row=row, column=2)
            third = worksheet.cell(row=row, column=3)

            if first.value in summary_titles and second.value in (None, ""):
                first.font = _SUMMARY_TITLE_FONT
                first.fill = _SUMMARY_TITLE_FILL
                worksheet.merge_cells(
                    start_row=row, start_column=1, end_row=row, end_column=3
                )
                first.alignment = _LEFT_TOP
                worksheet.row_dimensions[row].height = 20
                continue

            if first.value == "•" and second.value not in (None, ""):
                first.font = _TEXT_FONT
                second.alignment = _LEFT_TOP
                second.font = _TEXT_FONT
                third.alignment = _LEFT_TOP
                third.font = Font(
                    name="Calibri", size=10, bold=True, color="0A2540"
                )
                worksheet.row_dimensions[row].height = 28

    for row in range(data_start, last_row + 1):
        if worksheet.row_dimensions[row].height is None:
            worksheet.row_dimensions[row].height = 16


def _style_data_sheet(worksheet):

    last_row = worksheet.max_row
    last_col = worksheet.max_column

    if last_row == 0:
        return

    first_value = worksheet.cell(row=1, column=1).value
    is_section_header = (
        last_col == 1
        and isinstance(first_value, str)
        and worksheet.cell(row=1, column=2).value in (None, "")
    )

    if is_section_header:
        current_section = None
        table_first = None

        for row in range(1, last_row + 1):
            cell = worksheet.cell(row=row, column=1)
            value = cell.value

            if value in (None, ""):
                if table_first is not None and current_section is not None:
                    _apply_table_style(
                        worksheet, table_first, row - 1, last_col
                    )
                    table_first = None
                continue

            if table_first is None:
                if row == 1 or _looks_like_header(worksheet, row):
                    table_first = row
                    current_section = value
                    cell.font = _SECTION_FONT
                    cell.fill = _SECTION_FILL
                    cell.alignment = _LEFT
                else:
                    cell.font = _LABEL_FONT
                    cell.alignment = _LEFT
            else:
                if _looks_like_header(worksheet, row):
                    _apply_table_style(
                        worksheet, table_first, row - 1, last_col
                    )
                    table_first = row
                    current_section = value
                    cell.font = _SECTION_FONT
                    cell.fill = _SECTION_FILL
                    cell.alignment = _LEFT

        if table_first is not None:
            _apply_table_style(worksheet, table_first, last_row, last_col)

        worksheet.freeze_panes = "A2"
        return

    if _looks_like_header(worksheet, 1):
        _style_header_row(worksheet, 1, last_col)
        _apply_table_style(worksheet, 1, last_row, last_col)
        worksheet.freeze_panes = "A2"
    else:
        _apply_table_style(worksheet, 1, last_row, last_col)
        worksheet.freeze_panes = "A2"


def _looks_like_header(worksheet, row):

    first = worksheet.cell(row=row, column=1).value
    second = worksheet.cell(row=row, column=2).value

    if first in (None, "") or second in (None, ""):
        return False

    keywords = (
        "indicateur", "valeur", "identifiant", "localisation", "incidents",
        "interventions", "fournisseur", "ville", "agence", "region",
        "categorie", "mois", "total", "score", "etat", "statut", "dispo",
        "terminal", "libelle", "responsable", "filiale", "type",
        "level", "title", "message", "nbr",
    )

    text = "{} {}".format(str(first), str(second)).lower()
    return any(keyword in text for keyword in keywords)


class ExcelExportService:

    @staticmethod
    def export(statistics, filters=None, user=None):

        sheets = []

        claimed_sections = set()

        for name, builder, section in _SHEET_BUILDERS:
            if section == "filters":
                sheets.append((name, builder(statistics, filters)))
            elif name == "Résumé Exécutif":
                sheets.append((name, builder(statistics, filters, user)))
            else:
                sheets.append((name, builder(statistics, filters)))
            if section:
                claimed_sections.add(section)

        for section, data in statistics.items():
            if section in claimed_sections:
                continue
            sheets.append((section, _rows_from_value(data)))

        workbook = _build_workbook(sheets, user=user)

        buffer = BytesIO()
        workbook.save(buffer)

        response = HttpResponse(
            buffer.getvalue(), content_type=_CONTENT_TYPE_XLSX
        )
        filename = "ASIMS_Rapport_Statistiques_{0}.xlsx".format(
            timezone.now().strftime("%Y-%m-%d")
        )
        response["Content-Disposition"] = (
            'attachment; filename="{0}"'.format(filename)
        )

        return response
